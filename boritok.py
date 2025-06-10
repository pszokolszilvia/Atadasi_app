import os
import shutil
from openpyxl import load_workbook
import win32com.client as win32


def generate_boritok(excel_path, sablon_path, output_dir):
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    os.makedirs(output_dir)

    wb = load_workbook(excel_path)
    ws = wb.active

    relevant_rows = []
    for row in ws.iter_rows(min_row=2):
        sorszam, cim = row[0].value, row[1].value
        font = row[1].font
        if sorszam and cim and isinstance(cim, str):
            if not (font.strike or (font.color and font.color.rgb == "FFFF0000")):
                relevant_rows.append(f"{sorszam} {cim}".strip())

    # Eltávolítva: pythoncom.CoInitialize()
    word = win32.gencache.EnsureDispatch("Word.Application")
    word.Visible = False

    for i, cim in enumerate(relevant_rows, 1):
        doc = word.Documents.Open(os.path.abspath(sablon_path))
        try:
            rng = doc.Content
            find = rng.Find
            find.ClearFormatting()
            find.Replacement.ClearFormatting()
            find.Text = "[Ide írhat]"
            find.Replacement.Text = cim
            find.Execute(Replace=2)
        except Exception as e:
            print(f"Hiba {cim} cseréjénél: {e}")
        save_path = os.path.abspath(os.path.join(output_dir, f"borito_{i:03d}.docx"))
        doc.SaveAs(save_path)
        doc.Close()

    word.Quit()
